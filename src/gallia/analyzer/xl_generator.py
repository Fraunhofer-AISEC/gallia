# SPDX-FileCopyrightText: AISEC Pentesting Team
#
# SPDX-License-Identifier: Apache-2.0

"""
gallia-analyze EXCEL Generator module
"""

import json
from json.decoder import JSONDecodeError
from typing import Any, Dict, Tuple
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.exceptions import (
    InvalidFileException,
    SheetTitleException,
    WorkbookAlreadySaved,
    ReadOnlyWorkbookException,
)
import numpy as np
import pandas as pd
from pandas.core.indexing import IndexingError
from gallia.analyzer.operator import Operator
from gallia.analyzer.config import SrcPath, XlDesign, FAIL_CLS_CAP
from gallia.analyzer.failure import Failure
from gallia.analyzer.mode_config import LogMode, ScanMode
from gallia.analyzer.name_config import ColNm, ShtNm, CellCnt, KyNm
from gallia.services.uds.core.utils import g_repr


class ExcelGenerator(Operator):

    start_row: int = 1
    start_col: int = 1

    def __init__(self, path: str = "", log_mode: LogMode = LogMode.STD_OUT):
        Operator.__init__(self, path, log_mode)
        self.workbook: op.Workbook = op.Workbook()
        self.worksheet: Any
        self.load_color_code(SrcPath.err_src)

    def write_xl(
        self,
        entries_vec: np.ndarray,
        raw_df: pd.DataFrame,
        scan_mode: ScanMode,
        ecu_mode: int = -1,
    ) -> bool:
        self.load_ven_sess()
        try:
            if ecu_mode == -1:
                sum_sheet_name = "Summary"
                fail_sheet_suffix = ""
            else:
                sum_sheet_name = f"Summary({str(ecu_mode)})"
                fail_sheet_suffix = f"({str(ecu_mode)})"
            if ShtNm.init in self.workbook.sheetnames:
                self.workbook.remove(self.workbook[ShtNm.init])
            if scan_mode == ScanMode.SERV:
                if not self.add_sum_sheet_serv(raw_df, entries_vec, sum_sheet_name):
                    return False
                if not self.add_failure_sheet(raw_df, ScanMode.SERV, fail_sheet_suffix):
                    return False
            if scan_mode == ScanMode.IDEN:
                if not self.add_sum_sheet_iden(raw_df, entries_vec, sum_sheet_name):
                    return False
                if not self.add_failure_sheet(raw_df, ScanMode.IDEN, fail_sheet_suffix):
                    return False
            if len(self.workbook.worksheets) == 0:
                self.workbook.create_sheet(ShtNm.init)
        except (SheetTitleException, ReadOnlyWorkbookException) as exc:
            self.logger.error(f"generating EXCEL failed: {g_repr(exc)}")
            return False
        return True

    def save_close_xl(self, out_path: str) -> bool:
        try:
            self.workbook.save(out_path)
            self.workbook.close()
        except (InvalidFileException, WorkbookAlreadySaved) as exc:
            self.logger.error(f"saving EXCEL failed: {g_repr(exc)}")
            return False
        return True

    def add_sum_sheet_serv(
        self, raw_df: pd.DataFrame, entries_vec: np.ndarray, sheet_name: str = ""
    ) -> bool:
        """
        add summary sheet for scan_service to report EXCEL file.
        """
        if sheet_name == "":
            sheet_name = ShtNm.sum
        try:
            self.worksheet = self.workbook.create_sheet(sheet_name)
            ref_col = ColNm.serv
            dft_err_df = self.get_dft_err_df_from_raw(raw_df)
            cur_row, cur_col = self.sum_sheet_fill_origin(ScanMode.SERV)
            cur_row, cur_col = self.sum_sheet_fill_index(
                cur_row,
                cur_col,
                raw_df[raw_df["service"].isin(entries_vec)],
                ScanMode.SERV,
            )
            cur_row, cur_col = self.sum_sheet_fill_sess(cur_row, cur_col, dft_err_df)
            cur_row, cur_col = self.sum_sheet_fill_resp(
                cur_row,
                cur_col,
                dft_err_df,
                raw_df,
                ref_col,
                entries_vec,
                ScanMode.SERV,
            )
        except (KeyError, IndexError, AttributeError, SheetTitleException) as exc:
            self.logger.error(f"adding summary sheet failed: {g_repr(exc)}")
            return False
        return True

    def add_sum_sheet_iden(
        self, raw_df: pd.DataFrame, entries_vec: np.ndarray, sheet_name: str = ""
    ) -> bool:
        """
        add summary sheet for scan_identifier to report EXCEL file.
        """
        if sheet_name == "":
            sheet_name = ShtNm.sum
        try:
            self.worksheet = self.workbook.create_sheet(sheet_name)
            ref_col = ColNm.iden
            serv = np.unique(raw_df[ColNm.serv])[0]
            sbfn_vec = np.sort(np.unique(raw_df[ColNm.sbfn]))
            dft_err_df = self.get_dft_err_df_from_raw(raw_df)
            cur_row, cur_col = self.sum_sheet_fill_origin(ScanMode.IDEN, serv, sbfn_vec)
            cur_row, cur_col = self.sum_sheet_fill_index(
                cur_row,
                cur_col,
                raw_df[raw_df["identifier"].isin(entries_vec)],
                ScanMode.IDEN,
            )
            cur_row, cur_col = self.sum_sheet_fill_sess(cur_row, cur_col, dft_err_df)
            cur_row, cur_col = self.sum_sheet_fill_resp(
                cur_row,
                cur_col,
                dft_err_df,
                raw_df,
                ref_col,
                entries_vec,
                ScanMode.IDEN,
            )
        except (KeyError, IndexError, AttributeError, SheetTitleException) as exc:
            self.logger.error(f"adding summary sheet failed: {g_repr(exc)}")
            return False
        return True

    def sum_sheet_fill_origin(
        self,
        scan_mode: ScanMode,
        serv: int = 0,
        sbfn_vec: np.ndarray = np.array([]),
    ) -> Tuple[int, int]:
        """
        fill origin cell in summary sheet.
        """
        try:
            cur_row = self.start_row
            cur_col = self.start_col
            if scan_mode == ScanMode.SERV:
                header = CellCnt.serv
            if scan_mode == ScanMode.IDEN:
                header = self.get_code_text(serv, self.iso_serv_code_dict)
            self.worksheet.cell(cur_row, cur_col).value = header
            self.worksheet.cell(cur_row, cur_col).font = Font(name=XlDesign.font_index)
            self.set_cell_width(cur_col, XlDesign.dim_mid_wide)
            cur_row += 1
            self.worksheet.cell(cur_row, cur_col).value = CellCnt.default
            self.worksheet.cell(cur_row, cur_col).font = Font(name=XlDesign.font_index)
            cur_row += 1
            if sbfn_vec.size > 1:
                self.worksheet.cell(
                    self.start_row, self.start_col + 1
                ).value = CellCnt.sbfn
                self.worksheet.freeze_panes = self.worksheet.cell(
                    self.start_row + 1, self.start_col + 2
                ).coordinate
            else:
                self.worksheet.freeze_panes = self.worksheet.cell(
                    self.start_row + 1, self.start_col + 1
                ).coordinate
        except (KeyError, AttributeError) as exc:
            self.logger.error(
                f"filling origin cell of summary sheet failed: {g_repr(exc)}"
            )
            return self.start_row, self.start_col
        return cur_row, cur_col

    def sum_sheet_fill_index(
        self,
        cur_row: int,
        cur_col: int,
        entries_vec: pd.DataFrame,
        scan_mode: ScanMode,
    ) -> Tuple[int, int]:
        """
        fill index column in summary sheet.
        """
        has_sub_func = False
        has_id = False
        if scan_mode == ScanMode.SERV:
            entries_vec = entries_vec.drop_duplicates(["service"])
            has_id = True
        elif scan_mode == ScanMode.IDEN:
            entries_vec = entries_vec.drop_duplicates(["subfunc", "identifier"])
            if not entries_vec[entries_vec["subfunc"] != -1].empty:
                has_sub_func = True
            if not entries_vec[entries_vec["identifier"] != -1].empty:
                has_id = True
        else:
            raise NotImplementedError(f"ScanMode not supported: {scan_mode}")
        try:
            for _, row in entries_vec.iterrows():
                if scan_mode == ScanMode.SERV:
                    self.worksheet.cell(cur_row, cur_col).value = self.get_code_text(
                        row.service, self.iso_serv_code_dict
                    )
                    cur_row += 1
                if scan_mode == ScanMode.IDEN:
                    if row.identifier == -1:
                        index_name = CellCnt.no_ent
                    else:
                        index_name = f"0x{int(row.identifier):04X}"
                    self.worksheet.cell(cur_row, self.start_col).value = index_name
                    if row.subfunc != -1:
                        # service has subfunction and identifier
                        self.worksheet.cell(
                            cur_row, self.start_col + 1
                        ).value = row.subfunc
                    cur_row += 1

                self.worksheet.cell(cur_row, cur_col).font = Font(
                    name=XlDesign.font_index
                )
        except (KeyError, AttributeError) as exc:
            self.logger.error(
                f"filling index of summary sheet failed: {g_repr(exc)}"
            )
            return self.start_row, self.start_col + 1

        cur_col += int(has_id) + int(has_sub_func)
        cur_row = self.start_row
        return cur_row, cur_col

    def sum_sheet_fill_sess(
        self, cur_row: int, cur_col: int, dft_err_df: pd.DataFrame
    ) -> Tuple[int, int]:
        """
        fill top session row in summary sheet.
        """
        try:
            sess_vec = np.array(dft_err_df.columns)
            sess_num = 0
            for sess in sess_vec:
                dft_err = dft_err_df[sess][0]
                if dft_err == -1:
                    continue
                self.worksheet.cell(cur_row, cur_col).value = self.get_code_text(
                    sess, self.sess_code_dict
                )
                self.worksheet.cell(cur_row, cur_col).font = Font(
                    name=XlDesign.font_index
                )
                self.set_cell_width(cur_col, XlDesign.dim_wide)
                cur_row += 1
                self.worksheet.cell(cur_row, cur_col).value = self.get_code_text(
                    dft_err, self.iso_err_code_dict
                )
                self.worksheet.cell(cur_row, cur_col).font = Font(
                    name=XlDesign.font_value
                )
                self.fill_cell(cur_row, cur_col, dft_err)
                cur_row -= 1
                cur_col += 1
                sess_num += 1
            cur_col -= sess_num
            cur_row = self.start_row + 2
        except (KeyError, IndexingError, AttributeError) as exc:
            self.logger.error(
                f"filling top session row of summary sheet failed: {g_repr(exc)}"
            )
            return self.start_row + 1, self.start_col + 1
        return cur_row, cur_col

    def sum_sheet_fill_resp(
        self,
        cur_row: int,
        cur_col: int,
        dft_err_df: pd.DataFrame,
        raw_df: pd.DataFrame,
        ref_col: str,
        entries_vec: np.ndarray,
        scan_mode: ScanMode,
    ) -> Tuple[int, int]:
        """
        fill response field in summary sheet.
        """
        try:
            sess_vec = np.array(dft_err_df.columns)
            if scan_mode == ScanMode.SERV:
                sbfn_vec = np.arange(1)
            if scan_mode == ScanMode.IDEN:
                sbfn_vec = np.sort(np.unique(raw_df[ColNm.sbfn]))
            for sess in sess_vec:
                if dft_err_df[sess][0] == -1:
                    continue
                for entry in entries_vec:
                    for sbfn in sbfn_vec:
                        if scan_mode == ScanMode.SERV:
                            cond = (raw_df[ColNm.sess] == sess) & (
                                raw_df[ref_col] == entry
                            )
                        if scan_mode == ScanMode.IDEN:
                            cond = (
                                (raw_df[ColNm.sess] == sess)
                                & (raw_df[ref_col] == entry)
                                & (raw_df[ColNm.sbfn] == sbfn)
                            )
                        err_ser = raw_df.loc[cond, ColNm.resp].mode()
                        resp = self.get_code_text(
                            err_ser.iloc[-1], self.iso_err_code_dict
                        )
                        if err_ser.size == 1:
                            if err_ser[0] == 0:
                                search_id = int(raw_df[cond][ColNm.id].to_numpy()[0])
                                resp = (
                                    str(self.get_pos_res(search_id)) + "\n" + str(resp)
                                )
                        self.fill_cell(cur_row, cur_col, err_ser.iloc[-1])
                        self.worksheet.cell(cur_row, cur_col).value = resp
                        self.worksheet.cell(cur_row, cur_col).font = Font(
                            name=XlDesign.font_value
                        )
                        cur_row += 1
                cur_col += 1
                cur_row = self.start_row + 2
        except (KeyError, IndexingError, AttributeError) as exc:
            self.logger.error(
                f"filling response field of summary sheet failed: {g_repr(exc)}"
            )
            return self.start_row + 1, self.start_col + 1
        return cur_row, cur_col

    def add_failure_sheet(
        self, raw_df: pd.DataFrame, scan_mode: ScanMode, sheet_name_suffix: str = ""
    ) -> bool:
        """
        add failure(undocumented or missing) sheet to report EXCEL file.
        """
        if scan_mode == ScanMode.UNKNOWN:
            self.logger.error("adding summary sheet failed: scan mode unknown.")
            return False
        try:
            dft_err_df = self.get_dft_err_df_from_raw(raw_df)
            sess_vec = np.array(dft_err_df.columns)
            if scan_mode == ScanMode.SERV:
                fail_vec = np.array([Failure.UNDOC_SERV, Failure.MISS_SERV])
                width = XlDesign.dim_wide
            if scan_mode == ScanMode.IDEN:
                fail_vec = np.array([Failure.UNDOC_IDEN, Failure.MISS_IDEN])
                sbfn_vec = np.sort(np.unique(raw_df[ColNm.sbfn]))
                width = XlDesign.dim_middle
            cur_row = self.start_row
            cur_col = self.start_col
            sess_lu_vec = self.get_sess_lu()
            for fail in fail_vec:
                if fail in [Failure.UNDOC_SERV, Failure.UNDOC_IDEN]:
                    sheet_name = ShtNm.undoc
                if fail in [Failure.MISS_SERV, Failure.MISS_IDEN]:
                    sheet_name = ShtNm.miss
                self.worksheet = self.workbook.create_sheet(
                    f"{sheet_name}{sheet_name_suffix}"
                )
                self.worksheet.freeze_panes = self.worksheet.cell(
                    self.start_row + 1, self.start_col
                ).coordinate
                for sess in sess_vec:
                    self.set_cell_width(cur_col, width)
                    self.worksheet.cell(cur_row, cur_col).value = self.get_code_text(
                        sess, self.sess_code_dict
                    )
                    if dft_err_df[sess][0] == -1:
                        self.worksheet.cell(cur_row, cur_col).value = (
                            str(self.worksheet.cell(cur_row, cur_col).value)
                            + "\n"
                            + CellCnt.sess_unscn
                        )
                    if sess_lu_vec.size > 0:
                        if sess not in sess_lu_vec:
                            self.worksheet.cell(cur_row, cur_col).value = (
                                str(self.worksheet.cell(cur_row, cur_col).value)
                                + "\n"
                                + CellCnt.sess_undoc
                            )
                    self.set_cell_height(cur_row, XlDesign.dim_mid_wide)
                    self.worksheet.cell(cur_row, cur_col).alignment = Alignment(
                        horizontal="general", vertical="top"
                    )
                    self.worksheet.cell(cur_row, cur_col).font = Font(
                        name=XlDesign.font_index
                    )
                    cur_row += 1
                    cond = raw_df[ColNm.fail].apply(
                        lambda x, fl=fail: self.check_fail(x, fl)
                    ) & (raw_df[ColNm.sess] == sess)
                    if scan_mode == ScanMode.SERV:
                        serv_vec = np.sort(np.unique(raw_df.loc[cond, ColNm.serv]))
                        for serv in serv_vec:
                            self.worksheet.cell(
                                cur_row, cur_col
                            ).value = self.get_code_text(serv, self.iso_serv_code_dict)
                            self.worksheet.cell(cur_row, cur_col).font = Font(
                                name=XlDesign.font_value
                            )
                            cur_row += 1
                        cur_col += 1
                    if scan_mode == ScanMode.IDEN:
                        if sbfn_vec.size > 1:
                            raw_df[ColNm.combi] = list(
                                zip(raw_df[ColNm.iden], raw_df[ColNm.sbfn])
                            )
                            iden_sbfn_vec = np.sort(
                                np.unique(raw_df.loc[cond, ColNm.combi])
                            )
                            for iden_sbfn in iden_sbfn_vec:
                                iden = iden_sbfn[0]
                                sbfn = iden_sbfn[1]
                                if iden == -1:
                                    entry = CellCnt.no_ent
                                else:
                                    entry = f"0x{iden:04X} subfunc:{sbfn:02}"
                                self.worksheet.cell(cur_row, cur_col).value = entry
                                self.worksheet.cell(cur_row, cur_col).font = Font(
                                    name=XlDesign.font_value
                                )
                                cur_row += 1
                            cur_col += 1
                        else:
                            iden_vec = np.sort(np.unique(raw_df.loc[cond, ColNm.iden]))
                            for iden in iden_vec:
                                if iden == -1:
                                    entry = CellCnt.no_ent
                                else:
                                    entry = f"0x{iden:04X}"
                                self.worksheet.cell(cur_row, cur_col).value = entry
                                self.worksheet.cell(cur_row, cur_col).font = Font(
                                    name=XlDesign.font_value
                                )
                                cur_row += 1
                            cur_col += 1
                    cur_row = self.start_row
                cur_col = self.start_col
        except (KeyError, IndexingError, AttributeError, SheetTitleException) as exc:
            self.logger.error(
                f"adding failure summary sheets failed: {g_repr(exc)}"
            )
            return False
        return True

    def load_color_code(self, path: str) -> bool:
        """
        load color codes from JSON file
        """
        try:
            with open(path, encoding="utf8") as src_json:
                color_code_ls = json.load(src_json)
                self.color_code_dict = {
                    color_code[KyNm.err]: color_code[KyNm.rgb]
                    for color_code in color_code_ls
                }
        except (FileNotFoundError, KeyError, JSONDecodeError) as exc:
            self.logger.error(f"loading color codes failed: {g_repr(exc)}")
            return False
        return True

    def set_cell_width(self, col: int, width: int) -> bool:
        """
        set the cell width of given input column in the current EXCEL worksheet.
        """
        try:
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
        except (KeyError, AttributeError) as exc:
            self.logger.error(f"setting cell width failed: {g_repr(exc)}")
            return False
        return True

    def set_cell_height(self, row: int, height: int) -> bool:
        """
        set the cell height of given input row in the current EXCEL worksheet.
        """
        try:
            self.worksheet.row_dimensions[row].height = height
        except (KeyError, AttributeError) as exc:
            self.logger.error(f"setting cell height failed: {g_repr(exc)}")
            return False
        return True

    def fill_cell(self, row: int, col: int, error: int) -> bool:
        """
        fill a cell with color by given input error code.
        """
        try:
            self.worksheet.cell(row, col).fill = PatternFill(
                start_color=self.get_err_rgb(error),
                end_color=self.get_err_rgb(error),
                fill_type="solid",
            )
        except (KeyError, AttributeError) as exc:
            self.logger.error(f"filling cell failed: {g_repr(exc)}")
            return False
        return True

    def check_fail(self, fail: int, fail_class: Failure) -> bool:
        """
        check if given failure belongs to given faliure class.
        """
        return (fail // FAIL_CLS_CAP) == (fail_class // FAIL_CLS_CAP)

    def get_code_text(self, code: int, ref: Dict[int, str]) -> str:
        """
        get combined string of hex code and corresponding name
        with a given code and a given dictionary.
        """
        try:
            txt = ref[code]
        except KeyError:
            txt = "Unknown Code"
        if code in [-1, 0]:
            code_txt = f"{txt}"
        else:
            code_txt = f"0x{int(code):02X} {txt}"
        return code_txt

    def get_err_rgb(self, error: int) -> str:
        """
        get RGB color code string for an error response.
        """
        try:
            return "00" + self.color_code_dict[error]
        except KeyError:
            return "00FFFFFF"

    def get_gray_color(self, step: int) -> str:
        seed = ((step % 8) * 20) + 100
        return f"00{int(seed):02X}{int(seed):02X}{int(seed):02X}"
